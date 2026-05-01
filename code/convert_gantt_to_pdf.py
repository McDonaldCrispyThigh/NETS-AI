"""Convert Honors_Thesis_Gantt_Chart.xlsx to PDF using matplotlib."""
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import os, sys

# ── Load workbook ──────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook('docs/honors_registration/Honors_Thesis_Gantt_Chart.xlsx')
ws = wb.active

# ── Color map: xlsx 'AARRGGBB' → matplotlib hex ───────────────────────────────
COLORMAP = {
    '009DC3E6': '#9DC3E6',
    '00A9D18E': '#A9D18E',
    '00FFD966': '#FFD966',
    '00FF9999': '#FF9999',
}
GRAY_STRIPE = '#F5F5F5'
HEADER_COL  = '#2E4057'

# ── Read task rows 6-52 ───────────────────────────────────────────────────────
tasks = []
for rn in range(6, 53):
    wbs_val  = ws.cell(row=rn, column=1).value
    task_val = ws.cell(row=rn, column=2).value
    start    = ws.cell(row=rn, column=4).value
    due      = ws.cell(row=rn, column=5).value
    status   = ws.cell(row=rn, column=8).value

    bars = []
    for c in range(9, 41):
        cell = ws.cell(row=rn, column=c)
        fg   = cell.fill.fgColor
        rgb  = fg.rgb if fg.type == 'rgb' else None
        bars.append(COLORMAP.get(rgb, None))

    if wbs_val is not None or task_val is not None:
        def s(v):
            return str(v or '').encode('ascii', 'replace').decode()
        tasks.append({
            'wbs':    s(wbs_val),
            'task':   s(task_val),
            'start':  s(start),
            'due':    s(due),
            'status': s(status),
            'bars':   bars,
            'is_sec': wbs_val is None,
        })

# ── Figure layout ─────────────────────────────────────────────────────────────
N        = len(tasks)
WEEK_COLS = 32
ROW_H    = 0.30      # inches per row
LABEL_W  = 5.0       # inches: left label panel
BAR_W    = 8.5       # inches: bar panel
FIG_W    = LABEL_W + BAR_W
FIG_H    = max(12.0, N * ROW_H + 3.5)

fig = plt.figure(figsize=(FIG_W, FIG_H), facecolor='white')
fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
ax  = fig.add_axes([0, 0, 1, 1])
ax.set_xlim(0, FIG_W)
ax.set_ylim(0, FIG_H)
ax.axis('off')

# vertical positions
title_y    = FIG_H - 0.55
subtitle_y = FIG_H - 0.95
month_y    = FIG_H - 1.45
week_y     = FIG_H - 1.75
data_top   = FIG_H - 2.1

# left columns
col_task_x   = 0.15
col_wbs_w    = 0.45
col_task_w   = 2.80
col_start_x  = col_task_x + col_wbs_w + col_task_w + 0.05
col_start_w  = 0.65
col_due_x    = col_start_x + col_start_w + 0.05
col_due_w    = 0.65
col_status_x = col_due_x + col_due_w + 0.05

bar_left     = LABEL_W + 0.05
bar_cell_w   = (BAR_W - 0.1) / WEEK_COLS

# ── Title ─────────────────────────────────────────────────────────────────────
ax.text(FIG_W/2, title_y,
        'HONORS THESIS GANTT CHART',
        ha='center', va='center', fontsize=13, fontweight='bold', color=HEADER_COL)
ax.text(FIG_W/2, subtitle_y,
        'Evaluating Google Maps as a Pharmacy Establishment Data Source'
        '  |  Congyuan Zheng  |  CU Boulder  |  April 2026 - November 2026',
        ha='center', va='center', fontsize=8.5, color='#444444')

# ── Month headers ─────────────────────────────────────────────────────────────
months = ['April 2026','May 2026','June 2026','July 2026',
          'August 2026','September 2026','October 2026','November 2026']
month_bg = ['#E8F0F8','#F0F8F0','#FFF8E8','#F8E8E8',
            '#F8F0E8','#F0F8F8','#F8F8E8','#F8E8F0']
for mi, (mo, mc) in enumerate(zip(months, month_bg)):
    x0 = bar_left + mi * 4 * bar_cell_w
    x1 = x0 + 4 * bar_cell_w
    ax.add_patch(mpatches.FancyBboxPatch(
        (x0, month_y - 0.22), x1-x0, 0.28,
        boxstyle='square,pad=0', fc=mc, ec='#CCCCCC', lw=0.4))
    ax.text((x0+x1)/2, month_y - 0.08, mo,
            ha='center', va='center', fontsize=6.5, fontweight='bold', color=HEADER_COL)

for wi in range(WEEK_COLS):
    x0 = bar_left + wi * bar_cell_w
    ax.text(x0 + bar_cell_w/2, week_y - 0.08, f'W{(wi%4)+1}',
            ha='center', va='center', fontsize=5, color='#666666')

ax.axhline(data_top, color='#AAAAAA', lw=0.6)

# Left column headers
hdr_y = (month_y + data_top) / 2
for label, x in [('WBS', col_task_x),
                  ('Task', col_task_x + col_wbs_w + 0.05),
                  ('Start', col_start_x),
                  ('Due', col_due_x),
                  ('Status', col_status_x)]:
    ax.text(x, hdr_y, label, ha='left', va='center',
            fontsize=7, fontweight='bold', color=HEADER_COL)

# ── Task rows ─────────────────────────────────────────────────────────────────
STATUS_COLORS = {
    'Done': '#27AE60', 'In Progress': '#E67E22',
    'Pending': '#7F8C8D', 'DEADLINE': '#C0392B', 'MILESTONE': '#8E44AD',
}
SEC_BG = {
    'REGISTRATION': '#D9E8F5', 'WRITING': '#DFF0D8',
    'REVISION': '#FFF3CD', 'DEFENSE': '#F8D7DA',
}

for ri, t in enumerate(tasks):
    y_c   = data_top - (ri + 0.5) * ROW_H
    y_top = data_top - ri * ROW_H
    y_bot = y_top - ROW_H

    if t['is_sec']:
        key = next((k for k in SEC_BG if k in t['task'].upper()), None)
        fc  = SEC_BG.get(key, '#E8E8E8')
        ax.add_patch(mpatches.FancyBboxPatch(
            (0, y_bot), FIG_W, ROW_H,
            boxstyle='square,pad=0', fc=fc, ec='none'))
        ax.text(col_task_x, y_c, t['task'].strip(),
                ha='left', va='center', fontsize=7.5, fontweight='bold', color=HEADER_COL)
    else:
        stripe = GRAY_STRIPE if ri % 2 == 0 else 'white'
        ax.add_patch(mpatches.FancyBboxPatch(
            (0, y_bot), FIG_W, ROW_H,
            boxstyle='square,pad=0', fc=stripe, ec='none'))
        ax.text(col_task_x,                    y_c, t['wbs'],   ha='left', va='center', fontsize=6.5, color='#333333')
        ax.text(col_task_x + col_wbs_w + 0.05, y_c, t['task'],  ha='left', va='center', fontsize=6.5, color='#222222')
        ax.text(col_start_x,                   y_c, t['start'], ha='left', va='center', fontsize=6.5, color='#333333')
        ax.text(col_due_x,                     y_c, t['due'],   ha='left', va='center', fontsize=6.5, color='#333333')
        sc = STATUS_COLORS.get(t['status'], '#555555')
        ax.text(col_status_x, y_c, t['status'],
                ha='left', va='center', fontsize=6, color=sc, style='italic')

        # Gantt bars
        bar_h = ROW_H * 0.55
        bar_y  = y_c - bar_h / 2
        for wi, bar_color in enumerate(t['bars']):
            if bar_color:
                x0 = bar_left + wi * bar_cell_w
                ax.add_patch(mpatches.FancyBboxPatch(
                    (x0 + 0.01, bar_y), bar_cell_w - 0.02, bar_h,
                    boxstyle='square,pad=0', fc=bar_color, ec='none'))

    ax.axhline(y_bot, color='#DDDDDD', lw=0.3)

# Structural lines
ax.axvline(LABEL_W, color='#AAAAAA', lw=0.6,
           ymin=0, ymax=data_top/FIG_H)
for x0, y0, x1, y1 in [(0,0,FIG_W,0),(0,FIG_H,FIG_W,FIG_H),(0,0,0,FIG_H),(FIG_W,0,FIG_W,FIG_H)]:
    ax.plot([x0,x1],[y0,y1], color='#888888', lw=0.8)

# ── Legend ────────────────────────────────────────────────────────────────────
patches = [
    mpatches.Patch(fc='#9DC3E6', ec='none', label='Registration & Data'),
    mpatches.Patch(fc='#A9D18E', ec='none', label='Writing'),
    mpatches.Patch(fc='#FFD966', ec='none', label='Revision'),
    mpatches.Patch(fc='#FF9999', ec='none', label='Defense / Deadline'),
]
ax.legend(handles=patches, loc='lower left', bbox_to_anchor=(0.01, 0.005),
          ncol=4, fontsize=7, frameon=True, framealpha=0.9,
          facecolor='white', edgecolor='#BBBBBB')
ax.text(FIG_W * 0.5, 0.22,
        'KEY DEADLINES:  Apr 22 = Registration due  |  Nov 4 = Defense + copy due'
        '  |  Nov 6 = Faculty recs  |  Nov 10 = CU Scholar  |  Nov 13 = Honors designation',
        ha='center', va='center', fontsize=6.5, color='#333333', style='italic')

# ── Save ──────────────────────────────────────────────────────────────────────
out = 'docs/honors_registration/Honors_Thesis_Gantt_Chart.pdf'
plt.savefig(out, format='pdf', dpi=200, bbox_inches='tight', facecolor='white')
plt.close()
sz = os.path.getsize(out) / 1024
print(f'Saved: {out}  ({sz:.1f} KB)')
